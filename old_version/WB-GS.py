from selenium import webdriver
import os.path
from tkinter import *
from tkinter import ttk
import time
import random
from openpyxl import Workbook
import openpyxl
import datetime
import configparser
import threading
import sys
import os


# Vytvářím class pro hlavní ovládní botíka
class BotTop:
    # Kde mám ovladač a proxy server
    def __init__(self, driver, web):
        self.driver = driver  # Kde mám ovladač
        self.driver.get(web)

    # Log na stránku webgame
    def connect(self):
        self.driver.find_element_by_name('u').send_keys('www.webgame.cz')  # Odkaz na hru
        self.driver.find_element_by_xpath('//*[@id="go_btn"]/input[1]').click()

    # LogIn
    def login(self, jmeno, heslo):
        self.driver.find_element_by_name('login').send_keys(jmeno)
        self.driver.find_element_by_xpath('//*[@id="header"]/form/p[3]/input[1]').send_keys(heslo)
        self.driver.find_element_by_name('akce').click()

    # Proměná do které se mi importuje výběr z listu kde mám uložení menu
    def vyber(self, vyber):
        self.driver.find_element_by_partial_link_text(vyber).click()
    def click(self, xpath):
        self.driver.find_element_by_xpath(xpath).click()
    def berdata(self):
        driver = self.driver
        temp = driver.execute_script('return window.Highcharts.charts[0]''.series[0].options.data')
        print(temp)
        data = [item[1] for item in temp]
        print(data)

    def refresh(self):
        self.driver.refresh()

    def savedatanoexist(self):
        driver = self.driver
        # region Vytažené dat
        # K prodeji
        jidlo_pocetkprodeji = driver.find_element_by_xpath('//*[@id="wt_row_1"]/td[3]').text
        energie_pocetkprodeji = driver.find_element_by_xpath('//*[@id="wt_row_2"]/td[3]').text
        voj_pocetkprodeji = driver.find_element_by_xpath('//*[@id="wt_row_3"]/td[3]').text
        tanky_pocetkprodeji = driver.find_element_by_xpath('//*[@id="wt_row_4"]/td[3]').text
        stihy_pocetkprodeji = driver.find_element_by_xpath('//*[@id="wt_row_5"]/td[3]').text
        bunkry_pocetkprodeji = driver.find_element_by_xpath('//*[@id="wt_row_6"]/td[3]').text
        mechy_pocetkprodeji = driver.find_element_by_xpath('//*[@id="wt_row_6"]/td[3]').text
        # Cena
        jidlo_cena = driver.find_element_by_xpath('//*[@id="wt_row_1"]/td[5]').text
        jidlo_cena = jidlo_cena.split('\n')
        jidlo_cena = jidlo_cena[0]
        energie_cena = driver.find_element_by_xpath('//*[@id="wt_row_2"]/td[5]').text
        energie_cena = energie_cena.split('\n')
        energie_cena = energie_cena[0]
        voj_cena = driver.find_element_by_xpath('//*[@id="wt_row_3"]/td[5]').text
        voj_cena = voj_cena.split('\n')
        voj_cena = voj_cena[0]
        tanky_cena = driver.find_element_by_xpath('//*[@id="wt_row_4"]/td[5]').text
        tanky_cena = tanky_cena.split('\n')
        tanky_cena = tanky_cena[0]
        stihy_cena = driver.find_element_by_xpath('//*[@id="wt_row_5"]/td[5]').text
        stihy_cena = stihy_cena.split('\n')
        stihy_cena = stihy_cena[0]
        bunkry_cena = driver.find_element_by_xpath('//*[@id="wt_row_6"]/td[5]').text
        bunkry_cena = bunkry_cena.split('\n')
        bunkry_cena = bunkry_cena[0]
        mechy_cena = driver.find_element_by_xpath('//*[@id="wt_row_7"]/td[5]').text
        mechy_cena = mechy_cena.split('\n')
        mechy_cena = mechy_cena[0]
        # endregion
        # Proměné pro to kam zapsat hodnotu
        col = 0
        jidlo_row = int(jidlo_cena)
        ene_row = int(energie_cena)
        voj_row = int(voj_cena)
        tanky_row = int(tanky_cena)
        stihy_row = int(stihy_cena)
        bunkry_row = int(bunkry_cena)
        mech_row = int(mechy_cena)

        # Excel
        book = Workbook()
        sheet = book.active

        # region Zápis dat do sešitu
        # Jídlo
        sheet.cell(jidlo_row, col + 1).value = datetime.datetime.now()
        sheet.cell(jidlo_row, col + 2).value = jidlo_cena
        sheet.cell(jidlo_row, col + 3).value = jidlo_pocetkprodeji
        # Ene
        sheet.cell(ene_row, col + 4).value = datetime.datetime.now()
        sheet.cell(ene_row, col + 5).value = energie_cena
        sheet.cell(ene_row, col + 6).value = energie_pocetkprodeji
        # Vojáci
        sheet.cell(voj_row, col + 1).value = datetime.datetime.now()
        sheet.cell(voj_row, col + 2).value = voj_cena
        sheet.cell(voj_row, col + 3).value = voj_pocetkprodeji
        # tanky
        sheet.cell(tanky_row, col + 10).value = datetime.datetime.now()
        sheet.cell(tanky_row, col + 11).value = tanky_cena
        sheet.cell(tanky_row, col + 12).value = tanky_pocetkprodeji
        # Stíhy
        sheet.cell(stihy_row, col + 13).value = datetime.datetime.now()
        sheet.cell(stihy_row, col + 14).value = stihy_cena
        sheet.cell(stihy_row, col + 15).value = stihy_pocetkprodeji
        # Bunkry
        sheet.cell(bunkry_row, col + 16).value = datetime.datetime.now()
        sheet.cell(bunkry_row, col + 17).value = bunkry_cena
        sheet.cell(bunkry_row, col + 18).value = bunkry_pocetkprodeji
        # Mechy
        sheet.cell(mech_row, col + 19).value = datetime.datetime.now()
        sheet.cell(mech_row, col + 20).value = mechy_cena
        sheet.cell(mech_row, col + 21).value = mechy_pocetkprodeji
        # endregion

        # Uložení dat do sešitu
        book.save('data.xlsx')
        print('sešit neexistoval')

    def savedataexist(self):
        # osazení driveru
        driver = self.driver
        # region Vytáhnutí dat z tabulky světového trhu
        # K prodeji
        jidlo_pocetkprodeji = driver.find_element_by_xpath('//*[@id="wt_row_1"]/td[3]').text
        energie_pocetkprodeji = driver.find_element_by_xpath('//*[@id="wt_row_2"]/td[3]').text
        voj_pocetkprodeji = driver.find_element_by_xpath('//*[@id="wt_row_3"]/td[3]').text
        tanky_pocetkprodeji = driver.find_element_by_xpath('//*[@id="wt_row_4"]/td[3]').text
        stihy_pocetkprodeji = driver.find_element_by_xpath('//*[@id="wt_row_5"]/td[3]').text
        bunkry_pocetkprodeji = driver.find_element_by_xpath('//*[@id="wt_row_6"]/td[3]').text
        mechy_pocetkprodeji = driver.find_element_by_xpath('//*[@id="wt_row_7"]/td[3]').text

        # Cena
        jidlo_cena = driver.find_element_by_xpath('//*[@id="wt_row_1"]/td[5]').text
        jidlo_cena = jidlo_cena.split('\n')
        jidlo_cena = jidlo_cena[0]
        energie_cena = driver.find_element_by_xpath('//*[@id="wt_row_2"]/td[5]').text
        energie_cena = energie_cena.split('\n')
        energie_cena = energie_cena[0]
        voj_cena = driver.find_element_by_xpath('//*[@id="wt_row_3"]/td[5]').text
        voj_cena = voj_cena.split('\n')
        voj_cena = voj_cena[0]
        tanky_cena = driver.find_element_by_xpath('//*[@id="wt_row_4"]/td[5]').text
        tanky_cena = tanky_cena.split('\n')
        tanky_cena = tanky_cena[0]
        stihy_cena = driver.find_element_by_xpath('//*[@id="wt_row_5"]/td[5]').text
        stihy_cena = stihy_cena.split('\n')
        stihy_cena = stihy_cena[0]
        bunkry_cena = driver.find_element_by_xpath('//*[@id="wt_row_6"]/td[5]').text
        bunkry_cena = bunkry_cena.split('\n')
        bunkry_cena = bunkry_cena[0]
        mechy_cena = driver.find_element_by_xpath('//*[@id="wt_row_7"]/td[5]').text
        mechy_cena = mechy_cena.split('\n')
        mechy_cena = mechy_cena[0]
        # endregion

        # excel
        book = openpyxl.load_workbook('data.xlsx')
        sheet = book.active
        # Proměné pro to kam zapsat hodnotu
        col = 0
        jidlo_row = int(jidlo_cena)
        ene_row = int(energie_cena)
        voj_row = int(voj_cena)
        tanky_row = int(tanky_cena)
        stihy_row = int(stihy_cena)
        bunkry_row = int(bunkry_cena)
        mech_row = int(mechy_cena)

        # region Jídlo
        sheet.cell(jidlo_row, col + 1).value = datetime.datetime.now()
        sheet.cell(jidlo_row, col + 2).value = jidlo_cena
        sheet.cell(jidlo_row, col + 3).value = jidlo_pocetkprodeji
        jidlo_row_delete = jidlo_row - 1
        while jidlo_row_delete > 0:
            sheet.cell(jidlo_row_delete, col + 1).value = None
            sheet.cell(jidlo_row_delete, col + 2).value = None
            sheet.cell(jidlo_row_delete, col + 3).value = None
            jidlo_row_delete -= 1
        # endregion
        # region Energie
        sheet.cell(ene_row, col + 4).value = datetime.datetime.now()
        sheet.cell(ene_row, col + 5).value = energie_cena
        sheet.cell(ene_row, col + 6).value = energie_pocetkprodeji
        energie_row_delete = ene_row - 1
        while energie_row_delete > 0:
            sheet.cell(energie_row_delete, col + 4).value = None
            sheet.cell(energie_row_delete, col + 5).value = None
            sheet.cell(energie_row_delete, col + 6).value = None
            energie_row_delete -= 1
        # endregion
        # region Vojáci
        sheet.cell(voj_row, col + 7).value = datetime.datetime.now()
        sheet.cell(voj_row, col + 8).value = voj_cena
        sheet.cell(voj_row, col + 9).value = voj_pocetkprodeji

        voj_row_delete = voj_row - 1
        while voj_row_delete > 0:
            sheet.cell(voj_row_delete, col + 7).value = None
            sheet.cell(voj_row_delete, col + 8).value = None
            sheet.cell(voj_row_delete, col + 9).value = None
            voj_row_delete -= 1
        # endregion
        # region Tanky
        sheet.cell(tanky_row, col + 10).value = datetime.datetime.now()
        sheet.cell(tanky_row, col + 11).value = tanky_cena
        sheet.cell(tanky_row, col + 12).value = tanky_pocetkprodeji
        tanky_row_delete = tanky_row - 1
        while tanky_row_delete > 0:
            sheet.cell(tanky_row_delete, col + 10).value = None
            sheet.cell(tanky_row_delete, col + 11).value = None
            sheet.cell(tanky_row_delete, col + 12).value = None
            tanky_row_delete -= 1
        # endregion
        # region Stíhačky
        sheet.cell(stihy_row, col + 13).value = datetime.datetime.now()
        sheet.cell(stihy_row, col + 14).value = stihy_cena
        sheet.cell(stihy_row, col + 15).value = stihy_pocetkprodeji
        stihy_row_delete = stihy_row - 1
        while stihy_row_delete > 0:
            sheet.cell(stihy_row_delete, col + 13).value = None
            sheet.cell(stihy_row_delete, col + 14).value = None
            sheet.cell(stihy_row_delete, col + 15).value = None
            stihy_row_delete -= 1
        # endregion
        # region Bunkry
        sheet.cell(bunkry_row, col + 16).value = datetime.datetime.now()
        sheet.cell(bunkry_row, col + 17).value = bunkry_cena
        sheet.cell(bunkry_row, col + 18).value = bunkry_pocetkprodeji
        bunkry_row_delete = bunkry_row - 1
        while bunkry_row_delete > 0:
            sheet.cell(bunkry_row_delete, col + 16).value = None
            sheet.cell(bunkry_row_delete, col + 17).value = None
            sheet.cell(bunkry_row_delete, col + 18).value = None
            bunkry_row_delete -= 1
        # endregion
        # region Mechy
        sheet.cell(mech_row, col + 19).value = datetime.datetime.now()
        sheet.cell(mech_row, col + 20).value = mechy_cena
        sheet.cell(mech_row, col + 21).value = mechy_pocetkprodeji
        mech_row_delete = mech_row - 1
        while mech_row_delete > 0:
            sheet.cell(mech_row_delete, col + 19).value = None
            sheet.cell(mech_row_delete, col + 20).value = None
            sheet.cell(mech_row_delete, col + 21).value = None
            mech_row_delete -= 1
        # endregion

        print('Odmazáno to podtím')
        book.save('data.xlsx')

    def datasvetovytrh(self):
        print("Skenuju trh")
        driver = self.driver
        jidlo_pocetkprodeji = driver.find_element_by_xpath('//*[@id="wt_row_1"]/td[3]').text
        energie_pocetkprodeji = driver.find_element_by_xpath('//*[@id="wt_row_2"]/td[3]').text
        voj_pocetkprodeji = driver.find_element_by_xpath('//*[@id="wt_row_3"]/td[3]').text
        tanky_pocetkprodeji = driver.find_element_by_xpath('//*[@id="wt_row_4"]/td[3]').text
        stihy_pocetkprodeji = driver.find_element_by_xpath('//*[@id="wt_row_5"]/td[3]').text
        bunkry_pocetkprodeji = driver.find_element_by_xpath('//*[@id="wt_row_6"]/td[3]').text
        mechy_pocetkprodeji = driver.find_element_by_xpath('//*[@id="wt_row_6"]/td[3]').text
        # Cena
        jidlo_cena = driver.find_element_by_xpath('//*[@id="wt_row_1"]/td[5]').text
        jidlo_cena = jidlo_cena.split('\n')
        self.jidlo_cena = jidlo_cena[0]
        energie_cena = driver.find_element_by_xpath('//*[@id="wt_row_2"]/td[5]').text
        energie_cena = energie_cena.split('\n')
        self.energie_cena = energie_cena[0]
        voj_cena = driver.find_element_by_xpath('//*[@id="wt_row_3"]/td[5]').text
        voj_cena = voj_cena.split('\n')
        self.voj_cena = voj_cena[0]
        tanky_cena = driver.find_element_by_xpath('//*[@id="wt_row_4"]/td[5]').text
        tanky_cena = tanky_cena.split('\n')
        self.tanky_cena = tanky_cena[0]
        stihy_cena = driver.find_element_by_xpath('//*[@id="wt_row_5"]/td[5]').text
        stihy_cena = stihy_cena.split('\n')
        self.stihy_cena = stihy_cena[0]
        bunkry_cena = driver.find_element_by_xpath('//*[@id="wt_row_6"]/td[5]').text
        bunkry_cena = bunkry_cena.split('\n')
        self.bunkry_cena = bunkry_cena[0]
        mechy_cena = driver.find_element_by_xpath('//*[@id="wt_row_7"]/td[5]').text
        mechy_cena = mechy_cena.split('\n')
        self.mechy_cena = mechy_cena[0]

    def datadomacitrh(self):
        print("Skenuju trh")
        driver = self.driver
        jidlo_pocetkprodeji_dom = driver.find_element_by_xpath('//*[@class="vis_tbl"]/tbody/tr[2]/td[4]').text
        self.jidlo_pocetkprodeji_dom = jidlo_pocetkprodeji_dom.replace(" ", '')

        energie_pocetkprodeji_dom = driver.find_element_by_xpath('//*[@class="vis_tbl"]/tbody/tr[3]/td[4]').text
        self.energie_pocetkprodeji_dom = energie_pocetkprodeji_dom.replace(" ", '')

        voj_pocetkprodeji_dom = driver.find_element_by_xpath('//*[@class="vis_tbl"]/tbody/tr[4]/td[4]').text
        self.voj_pocetkprodeji_dom = voj_pocetkprodeji_dom.replace(" ", '')

        tanky_pocetkprodeji_dom = driver.find_element_by_xpath('//*[@class="vis_tbl"]/tbody/tr[5]/td[4]').text
        self.tanky_pocetkprodeji_dom = tanky_pocetkprodeji_dom.replace(" ", '')

        stihy_pocetkprodeji_dom = driver.find_element_by_xpath('//*[@class="vis_tbl"]/tbody/tr[6]/td[4]').text
        self.stihy_pocetkprodeji_dom = stihy_pocetkprodeji_dom.replace(" ", '')

        bunkry_pocetkprodeji_dom = driver.find_element_by_xpath('//*[@class="vis_tbl"]/tbody/tr[7]/td[4]').text
        self.bunkry_pocetkprodeji_dom = bunkry_pocetkprodeji_dom.replace(" ", '')

        mechy_pocetkprodeji_dom = driver.find_element_by_xpath('//*[@class="vis_tbl"]/tbody/tr[8]/td[4]').text
        self.mechy_pocetkprodeji_dom = mechy_pocetkprodeji_dom.replace(" ", '')

        # Cena
        jidlo_cena = driver.find_element_by_xpath('//*[@class="vis_tbl"]/tbody/tr[2]/td[6]').text
        jidlo_cena = jidlo_cena.replace('$', '')
        self.jidlo_cena_dom = jidlo_cena.replace(" ", '')
        energie_cena = driver.find_element_by_xpath('//*[@class="vis_tbl"]/tbody/tr[3]/td[6]').text
        self.energie_cena_dom = energie_cena.replace('$', '')
        voj_cena = driver.find_element_by_xpath('//*[@class="vis_tbl"]/tbody/tr[4]/td[6]').text
        self.voj_cena_dom = voj_cena.replace('$', '')
        tanky_cena = driver.find_element_by_xpath('//*[@class="vis_tbl"]/tbody/tr[5]/td[6]').text
        self.tanky_cena_dom = tanky_cena.replace('$', '')
        stihy_cena = driver.find_element_by_xpath('//*[@class="vis_tbl"]/tbody/tr[6]/td[6]').text
        self.stihy_cena_dom = stihy_cena.replace('$', '')
        bunkry_cena = driver.find_element_by_xpath('//*[@class="vis_tbl"]/tbody/tr[7]/td[6]').text
        self.bunkry_cena_dom = bunkry_cena.replace('$', '')
        mechy_cena = driver.find_element_by_xpath('//*[@class="vis_tbl"]/tbody/tr[8]/td[6]').text
        self.mechy_cena_dom = mechy_cena.replace('$', '')

    def vypocetdomtosv(self):
        print("Pocitam trh")
        jidlo_pocetknakupu_dom = self.jidlo_pocetkprodeji_dom
        energie_pocetknakupu_dom = self.energie_pocetkprodeji_dom
        voj_pocetknakupu_dom = self.voj_pocetkprodeji_dom
        tanky_pocetknakupu_dom = self.tanky_pocetkprodeji_dom
        stihy_pocetknakupu_dom = self.stihy_pocetkprodeji_dom
        bunkry_pocetknakupu_dom = self.bunkry_pocetkprodeji_dom
        mechy_pocetknakupu_dom = self.mechy_pocetkprodeji_dom
        # Cena dom trh
        jidlo_cena_dom = self.jidlo_cena_dom
        print(jidlo_cena_dom)
        energie_cena_dom = self.energie_cena_dom
        print(energie_cena_dom )
        voj_cena_dom = self.voj_cena_dom
        print(voj_cena_dom)
        tanky_cena_dom = self.tanky_cena_dom
        print(tanky_cena_dom)
        stihy_cena_dom = self.stihy_cena_dom
        print(stihy_cena_dom)
        bunkry_cen_doma = self.bunkry_cena_dom
        print(bunkry_cen_doma)
        mechy_cena_dom = self.mechy_cena_dom
        print(mechy_cena_dom)
        # Cena sv trh
        print('uložení světového trhu cen')
        jidlo_cena_sv = self.jidlo_cena
        energie_cena_sv = self.energie_cena
        voj_cena_sv = self.voj_cena
        tanky_cena_sv = self.tanky_cena
        stihy_cena_sv = self.stihy_cena
        bunkry_cen_sv = self.bunkry_cena
        mechy_cena_sv = self.mechy_cena
        print('jdu počítat jídlo')
        self.vydelek_jidlo = ((int(jidlo_pocetknakupu_dom) * int(jidlo_cena_sv)) * 0.90) - (int(jidlo_pocetknakupu_dom) * int(jidlo_cena_dom))
        print('ene')
        self.vydelek_ene = ((int(energie_pocetknakupu_dom) * int(energie_cena_sv)) * 0.90) - (int(energie_pocetknakupu_dom) * int(energie_cena_dom))
        print('vojáci')
        self.vydelek_voj = ((int(voj_pocetknakupu_dom) * int(voj_cena_sv)) * 0.94) - (int(voj_pocetknakupu_dom) * int(voj_cena_dom)) #KK
        self.vydelek_tanky = ((int(tanky_pocetknakupu_dom) * int(tanky_cena_sv)) * 0.94) - (int(tanky_pocetknakupu_dom) * int(tanky_cena_dom))
        self.vydelek_stihy = ((int(stihy_pocetknakupu_dom) * int(stihy_cena_sv)) * 0.94) - (int(stihy_pocetknakupu_dom) * int(stihy_cena_dom))
        self.vydelek_bunkry = ((int(bunkry_pocetknakupu_dom) * int(bunkry_cen_sv)) * 0.94) - (int(bunkry_pocetknakupu_dom) * int(bunkry_cen_doma))
        self.vydelek_mechy = ((int(mechy_pocetknakupu_dom) * int(mechy_cena_sv)) * 0.94) - (int(mechy_pocetknakupu_dom) * int(mechy_cena_dom))

    def prodejdomtosv(self):
        driver = self.driver
        print("Start vepsání")
        list_kontrola = [self.vydelek_jidlo, self.vydelek_ene, self.vydelek_voj, self.vydelek_tanky, self.vydelek_stihy, self.vydelek_bunkry, self.vydelek_mechy]
        print(list_kontrola)
        penize = driver.find_element_by_xpath('//*[@class="rdata"]').text
        penize = penize.split()
        print(penize[1])

        pb = 0
        while pb < 7:
            list_cen = [self.vydelek_jidlo, self.vydelek_ene, self.vydelek_voj, self.vydelek_tanky, self.vydelek_stihy, self.vydelek_bunkry, self.vydelek_mechy]
            list_cen.sort(reverse=True)
            tr = 0
            print("To kolik vydělám:")
            print(list_cen)
            if list_cen[pb] == self.vydelek_jidlo:
                tr = "2"
                pocet = self.jidlo_pocetkprodeji_dom
                text_prodej = "jídlo"
            elif list_cen[pb] == self.vydelek_ene:
                tr = "3"
                pocet = self.energie_pocetkprodeji_dom
                text_prodej = "Energii"
            elif list_cen[pb] == self.vydelek_voj:
                tr = "4"
                pocet = self.voj_pocetkprodeji_dom
                text_prodej = "Vojáky"
            elif list_cen[pb] == self.vydelek_tanky:
                tr = "5"
                pocet = self.tanky_pocetkprodeji_dom
                text_prodej = "Tanky"
            elif list_cen[pb] == self.vydelek_stihy:
                tr = "6"
                pocet = self.stihy_pocetkprodeji_dom

                text_prodej = "Stíhy"
            elif list_cen[pb] == self.vydelek_bunkry:
                tr = "7"
                pocet = self.bunkry_pocetkprodeji_dom
                text_prodej = "Bunkry"
            elif list_cen[pb] == self.vydelek_mechy:
                tr = "8"
                pocet = self.mechy_pocetkprodeji_dom
                text_prodej = "Mechy"
            else:
                break
            if list_cen[pb] > 100:
                if pocet >= 500:
                    driver.find_element_by_xpath('//*[@class="vis_tbl"]/tbody/tr['+tr+']/td[8]/input').send_keys(pocet)
                elif pocet >= 2000:
                    driver.find_element_by_xpath('//*[@class="vis_tbl"]/tbody/tr[' + tr + ']/td[8]/input').send_keys(pocet)
                else:
                    print("Nevyplatí se mi prodávat, kvuli množství:"+ text_prodej)
            else:
                print("Nevyplatí se mi prodávat" + text_prodej + ", protože bych vydělal méně než 100$")
            pb += 1
            print("Konec cyklu")


    def closedriver(self):
        self.driver.quit()

# Def ze kterého spouštím nějaký chod programu, uvnitř jsou vložené instance, metody z třídy která je nad tím
def start():
    print(run)
    print(prohlizec.get())
    print(chdrive.get())
    print(login.get())
    print(heslo.get())
    # Zda chceš ppužít proxy nebo ne
    if "Chrome" == prohlizec.get():
        if proxy.get() == 1:
            # s proxy
            pripojeni = BotTop(webdriver.Chrome(chdrive.get()),"https://www.free-proxy.com/")  # Cesta k driveru 'C:\\bin\\chromedriver'
            pripojeni.connect()
            time.sleep(3)
        else:
            # bez proxy
            pripojeni = BotTop(webdriver.Chrome(chdrive.get()),"https://www.webgame.cz/")  # Cesta k driveru 'C:\\bin\\chromedriver'
    elif prohlizec.get() == "Firefox":
        if proxy.get() == 1:
            # s proxy
            pripojeni = BotTop(webdriver.Firefox(chdrive.get()),"https://www.free-proxy.com/")  # Cesta k driveru 'C:\\bin\\chromedriver'
            pripojeni.connect()
            time.sleep(3)
        else:
            # bez proxy
            pripojeni = BotTop(webdriver.Firefox(executable_path=chdrive.get()),"https://www.webgame.cz/")  # Cesta k driveru 'C:\\bin\\chromedriver'
    elif prohlizec == "Internet Explorer":
        if proxy.get() == 1:
            # s proxy
            pripojeni = BotTop(webdriver.Ie(chdrive.get()),"https://www.free-proxy.com/")  # Cesta k driveru 'C:\\bin\\chromedriver'
            pripojeni.connect()
            time.sleep(3)
        else:
            # bez proxy
            pripojeni = BotTop(webdriver.Ie(chdrive.get()),"https://www.webgame.cz/")  # Cesta k driveru 'C:\\bin\\chromedriver'
    else:
        print("Nevybral jsi prohlížeč")
    time.sleep(4)
    # vložení log inu a heslo do formuláře a odeslání
    pripojeni.login(login.get(), heslo.get())
    time.sleep(3)
    # random výběr klikání nebo jen refreshe na Světovém trhu
    if refresh.get() == 0:
        while run == True:
            # vbere náhodně nějaký odkaz (54% šance na sv, zbytek se pak dělí po pár procentech)
            rch = random.choice(listsodkazem)
            pripojeni.vyber(rch)  # kliknutí na vybraný odkaz
            try:
                if rch == 'Sv':
                    # pokud jsi klikl na Sv tak proces usne a ověří si jestli excel existuje nebo ne
                    x = random.randrange(2, 30)
                    print('Usnu na:' + str(x) + 'sekund')

                    dataexist = os.path.exists('data.xlsx')
                    if not dataexist:
                        pripojeni.savedatanoexist()  # Když excel neexsituje tak mi vytvoří nový a zapíše do něj hodnoty
                    else:
                        pripojeni.savedataexist()  # Když excel exsituje tak zapíše do něj hodnoty

                    time.sleep(x)
                else:
                    sleeptime = random.randrange(1, 15)
                    print('Nevybral jsem Svtrh spím na ' + str(sleeptime) + 'sekund')  # Kód zastaví a pak znova opakuje
                    time.sleep(sleeptime)
            except:
                print('Něco se pokazilo!!')
        pripojeni.closedriver()
    else:  # Refresh jen světového trhu
        while run == True:
            pripojeni.vyber('Sv')
            txcs = random.randrange(2, 30)  # Random refresh od 2 do 30 sekund
            time.sleep(txcs)
            pripojeni.refresh()  # poriběhne refresh
            dataexist = os.path.exists('data.xlsx')
            try:
                if not dataexist:
                    pripojeni.savedatanoexist()
                else:
                    pripojeni.savedataexist()
            except:
                print('Něco se pokazilo!!!')
        pripojeni.closedriver()
def konfigurace():
    config = configparser.ConfigParser()
    config['Data'] = {'Path': chdrive.get(), 'LogIn': login.get(), 'Password': heslo.get()}
    config.write(open('config.ini', 'w'))
    print('Done')
def configdelete():
    heslo.delete(0, END)
    login.delete(0, END)
    chdrive.delete(0, END)
    print('Přepsáno')
def konec():
    global run
    run = False
    print('Konec')
def runtrue():
    global run
    run = True
def t_dom():
    print(run)
    print(prohlizec.get())
    print(chdrive.get())
    print(login.get())
    print(heslo.get())
    # Zda chceš ppužít proxy nebo ne
    if "Chrome" == prohlizec.get():
        if proxy.get() == 1:
            # s proxy
            pripojeni = BotTop(webdriver.Chrome(chdrive.get()),"https://www.free-proxy.com/")  # Cesta k driveru 'C:\\bin\\chromedriver'
            pripojeni.connect()
            time.sleep(3)
        else:
            # bez proxy
            pripojeni = BotTop(webdriver.Chrome(chdrive.get()),"https://www.webgame.cz/")  # Cesta k driveru 'C:\\bin\\chromedriver'
    elif prohlizec.get() == "Firefox":
        if proxy.get() == 1:
            # s proxy
            pripojeni = BotTop(webdriver.Firefox(chdrive.get()),"https://www.free-proxy.com/")  # Cesta k driveru 'C:\\bin\\chromedriver'
            pripojeni.connect()
            time.sleep(3)
        else:
            # bez proxy
            pripojeni = BotTop(webdriver.Firefox(executable_path=chdrive.get()),"https://www.webgame.cz/")  # Cesta k driveru 'C:\\bin\\chromedriver'
    elif prohlizec == "Internet Explorer":
        if proxy.get() == 1:
            # s proxy
            pripojeni = BotTop(webdriver.Ie(chdrive.get()),"https://www.free-proxy.com/")  # Cesta k driveru 'C:\\bin\\chromedriver'
            pripojeni.connect()
            time.sleep(3)
        else:
            # bez proxy
            pripojeni = BotTop(webdriver.Ie(chdrive.get()),"https://www.webgame.cz/")  # Cesta k driveru 'C:\\bin\\chromedriver'
    else:
        print("Nevybral jsi prohlížeč")
    time.sleep(4)
    # vložení log inu a heslo do formuláře a odeslání
    pripojeni.login(login.get(), heslo.get())
    time.sleep(3)
    pripojeni.vyber('Sv')
    pripojeni.click('//*[@id="uMenu"]/ul/li[1]/a[3]')
    pripojeni.berdata()

    #pripojeni.closedriver()



# Nastavení proměných
run = True
odkaz = ['Sv', 'Det', 'Arch', 'Dom', 'Roz']  # list s možnými odkazy
listsodkazem = ['Sv', random.choice(odkaz)]

config = configparser.ConfigParser()  # nastavení čtení configu
config.read('config.ini')

# regionNastavení proměných ppro výpln textu při zapnutí
configexist = os.path.exists('config.ini')
if not configexist:
    heslo_insert = ''  # Nastavuju prázdnou hodnotu
    login_insert = ''
    driver_insert = ''
else:
    heslo_insert = config['Data']['password']  # Nastavuju čtení hodnot z configu
    login_insert = config['Data']['login']
    driver_insert = config['Data']['path']
# endregion
# region Win Okno na spouštění appky
# Formulář window
window = Tk()
window.title("Webgame Trh")
window.geometry('300x400')
# region Drive/LogiN část
# Kde máš drive k chrome
prohlizec_text = Label(window, text="Vyber si prohlížeč:").place(relx=.2, rely=.1, anchor=S)
prohlizec = ttk.Combobox(window, values=("Chrome", "Internet Explorer", "Firefox"), width=10)
prohlizec.place(relx=.6, rely=.1, anchor=S)

chdrive_text = Label(window, text="Drive:").place(relx=.2, rely=.17, anchor=S)
chdrive = Entry(window, text='Ahoj')
chdrive.insert(1, driver_insert)
chdrive.place(relx=.6, rely=.17, anchor=S)
# Log In Input
login_text = Label(window, text="LogIn:").place(relx=.2, rely=.24, anchor=S)
login = Entry(window)
login.insert(1, login_insert)
login.place(relx=.6, rely=.24, anchor=S)
# Heslo
ahoj = 'watafaka'
heslo_text = Label(window, text="Heslo:").place(relx=.2, rely=.32, anchor=S)
heslo = Entry(window)
heslo.insert(1, heslo_insert)
heslo.place(relx=.6, rely=.32, anchor=S)
# endregion
# region Check buttons
refresh = IntVar()
Checkbutton(window, text='Refresh na Světovém trhu', variable=refresh).place(relx=.5, rely=.39, anchor=S)
proxy = IntVar()
Checkbutton(window, text='Proxy', variable=proxy).place(relx=.5, rely=.46, anchor=S)


# endregion
# region Tlačítka

t = threading.Thread(target=start)
t_dom = threading.Thread(target=t_dom)

Start = Button(window, text="Start", command=lambda: [runtrue(), t.start()])
Start.place(relx=.2, rely=.9, anchor=S)

p_domtrhButton = Button(window, text="Přeprodej dom. trhu", command=lambda: [runtrue(), t_dom.start()])
p_domtrhButton.place(relx=.2, rely=.8, anchor=S)
ConfigButton = Button(window, text="Vytvoř config", command=konfigurace)
ConfigButton.place(relx=.7, rely=.7, anchor=S)

ConfigButtondelete = Button(window, text="Vymaž buňky", command=lambda:configdelete())
ConfigButtondelete.place(relx=.7, rely=.8, anchor=S)

Konecbutton = Button(window, text="Ukonči zápis", command=lambda:konec())
Konecbutton.place(relx=.7, rely=.9, anchor=S)
# endregion
# endregion

# Window aktualizaace

window.mainloop()
# endregion
